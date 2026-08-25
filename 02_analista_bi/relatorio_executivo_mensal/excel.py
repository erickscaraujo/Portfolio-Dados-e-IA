"""Exportacao do relatorio executivo em Excel com formatacao profissional via openpyxl."""

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AZUL = PatternFill("solid", fgColor="1F4E79")
AMARELO = PatternFill("solid", fgColor="FFF2CC")
VERMELHO = Font(color="C00000")
FONTE_BRANCA = Font(color="FFFFFF", bold=True)


def _cabecalho(ws, df: pd.DataFrame, linha: int = 1) -> None:
    for coluna, nome in enumerate(df.columns, start=1):
        celula = ws.cell(row=linha, column=coluna, value=nome)
        celula.fill = AZUL
        celula.font = FONTE_BRANCA
        celula.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)


def _ajustar_larguras(ws, df: pd.DataFrame) -> None:
    for coluna, nome in enumerate(df.columns, start=1):
        maior = max(
            len(str(nome)),
            *(len(f"{v:,.0f}") if isinstance(v, float) else len(str(v)) for v in df[nome]),
        )
        ws.column_dimensions[get_column_letter(coluna)].width = min(maior + 4, 32)


def exportar(
    kpi_df: pd.DataFrame,
    detalhe_df: pd.DataFrame,
    caminho: str = "outputs/relatorio_executivo.xlsx",
) -> str:
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        kpi_df.to_excel(writer, sheet_name="Resumo", index=False)
        detalhe_df.to_excel(writer, sheet_name="Detalhe Mensal", index=False)

        ws_resumo = writer.sheets["Resumo"]
        _cabecalho(ws_resumo, kpi_df)
        # realca KPIs abaixo da meta para leitura rapida da diretoria
        coluna_status = list(kpi_df.columns).index("status") + 1
        for linha in range(2, len(kpi_df) + 2):
            celula = ws_resumo.cell(row=linha, column=coluna_status)
            if celula.value == "abaixo da meta":
                celula.fill = AMARELO
                celula.font = VERMELHO
        _ajustar_larguras(ws_resumo, kpi_df)

        ws_detalhe = writer.sheets["Detalhe Mensal"]
        _cabecalho(ws_detalhe, detalhe_df)
        ws_detalhe.auto_filter.ref = ws_detalhe.dimensions
        _ajustar_larguras(ws_detalhe, detalhe_df)

    return caminho
